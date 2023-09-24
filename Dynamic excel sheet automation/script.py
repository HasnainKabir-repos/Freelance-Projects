import openpyxl
import sys
import re

shipping_data = []

row_index = 1

try:
    
    print("Processing Please wait...")
    
    book = openpyxl.load_workbook('ClientIndexFGT.xlsx')
    sheet = book['Sheet1']
    
    while row_index <= sheet.max_row:

        row = sheet[row_index]
        B = row[1].value   

        if row_index >= 2299:
            break

        if len(str(B)) > 2 and row[0] != 'P':

            names = []
            locations = []
            FTS = ""

            names.append(B)

            if row[4].value == "FTS-1":
                FTS = "FTS-1"
            elif row[4].value == "FTS-2":
                FTS = "FTS-2"

            new_row_index = row_index + 1

            while new_row_index <= sheet.max_row:

                inner_row = sheet[new_row_index]
                new_B = inner_row[1].value

                if len(str(new_B)) > 2 and inner_row[0].value != 'P':

                    names.append(new_B)

                    if inner_row[4].value == "FTS-1":
                        FTS = "FTS-1"
                    elif inner_row[4].value == "FTS-2":
                        FTS = "FTS-2"
                else:
                    break

                new_row_index += 1

            while new_row_index <= sheet.max_row:

                resume_row = sheet[new_row_index]

                if len(str(resume_row[1].value)) <= 2 and resume_row[0].value == 'P':
                    locations.append(resume_row[2].value)
                else:
                    break

                new_row_index += 1

            data = {"Names" : names, "Locations" : locations, "FTS" : FTS}

            shipping_data.append(data)

        row_index = new_row_index
        
    print("Data Fetched Successfully") 
    
except FileNotFoundError:
    
    print('File not found, please keep the excel file in the same folder as this script')
    sys.exit(1)
    
except openpyxl.utils.exceptions.InvalidFileException:
    
    print('Invalid Excel file format. Please check the file format.')
    sys.exit(1)
    
except Exception as e:
    
    print('An error occurred: '+ str(e))
    sys.exit(1)


try:
    while True:
        
        print(f"-------------------------------Search--------------------------")
        
        location = str(input("Write the name you want to search for: "))
        fts = str(input("Enter the FTS number. (Leave empty if you want to see data for all FTS) "))

        location = location.upper()

        if fts:
            fts = fts[-1]

        for data in shipping_data:
            values = data['Locations']

            for i in values:
                if re.search(location, i, re.IGNORECASE):

                    if (fts and fts == '1' and data['FTS'] == 'FTS-1') or (fts and fts == '2' and data['FTS'] == 'FTS-2') or not fts:
                        print(f"{data['Names']} {i} {data['FTS']}")

        choice = input("Enter e to end or press enter to continue: ")
        
        if choice == "e" or choice == "E":
            break
            
except KeyboardInterrupt:
    print("Operation interrupted by user")
except Exception as e:
    print("An error occurred: " + str(e))