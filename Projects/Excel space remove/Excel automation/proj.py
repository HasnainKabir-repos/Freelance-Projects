from openpyxl import Workbook, load_workbook

wb = load_workbook('output.xlsx')
ws = wb.active

#Changing value of A1 cell
ws['A1'] = "No."

#To save the workbook
wb.save('output.xlsx')